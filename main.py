#!/usr/bin/env python3
# coding=utf-8

import contextlib
import mmap
import os
import sys
import glob
import datetime

import openpyxl as oxl
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from Evtx.Evtx import FileHeader
from Evtx.Views import evtx_file_xml_view
from lxml import etree


class AnalysisEvtx:
    def __init__(self, filePath, outputDir):
        self.fileName = os.path.splitext(os.path.basename(filePath))[0]
        self.requireTagList = [
            "Provider.Name",
            "Provider.Guid",
            "EventID",
            "Level",
            "TimeCreated.SystemTime",
            "EventRecordID",
            "Execution.ProcessID",
            "Execution.ThreadID",
            "Channel",
            "Computer",
            "ProcessID",
            "Application",
            "Direction",
            "SourceAddress",
            "SourcePort",
            "DestAddress",
            "DestPort",
            "Protocol",
            "RemoteUserID",
            "RemoteMachineID",
            "Security.UserID",
            "QueryName",
            "EventSourceName",
            "Data",
            "Binary",
        ]
        self.filePath = filePath
        self.outputDir = outputDir
        self.startAnalysis()

    def convertEvtxToXmlList(self, filePath):
        resultDicList = []
        xmlList = []
        with open(filePath, "rb") as f:
            with contextlib.closing(
                mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ)
            ) as buf:
                fh = FileHeader(buf, 0)
                nowCount = 1
                for xml, record in evtx_file_xml_view(fh):
                    print(f"\r解析第{nowCount}条事件", end="")
                    reDic = self.analysisXml(xml)
                    xmlList.append(xml)
                    resultDicList.append(reDic)
                    nowCount += 1
            print("<--解析完成\n")
        return xmlList, resultDicList

    def analysisXml(self, xmlStr):
        reDic = {tag: "" for tag in self.requireTagList}
        xmlObj = etree.XML(xmlStr)
        sysObj = xmlObj[0]
        dataObj = xmlObj[1]

        for item in sysObj:
            nowTag = "}".join(item.tag.split("}")[1:])
            if item.text is None:
                for tmpName, tmpValue in item.items():
                    tmpNowTag = nowTag + "." + tmpName
                    if self.checkIfRequire(tmpNowTag, self.requireTagList):
                        reDic = self.writeToDic(reDic, tmpNowTag, tmpValue)
            else:
                if self.checkIfRequire(nowTag, self.requireTagList):
                    reDic = self.writeToDic(reDic, nowTag, item.text)

        for item in dataObj:
            tmpAttrDic = {key: value for key, value in item.items()}
            nowValue = item.text
            if "Name" in tmpAttrDic.keys():
                nowTag = tmpAttrDic["Name"]
                if not self.checkIfRequire(nowTag, self.requireTagList):
                    nowTag = "Data"
                    nowValue = (
                        tmpAttrDic["Name"]
                        + ":"
                        + ("" if item.text is None else item.text)
                    )
            else:
                nowTag = "Data"
            reDic = self.writeToDic(reDic, nowTag, nowValue)
        return reDic

    def checkIfRequire(self, tagName, requireList):
        return tagName in requireList

    def writeToDic(self, aimDic, key, value):
        if value is None:
            value = ""
        if aimDic[key] == "":
            aimDic[key] = value
        else:
            aimDic[key] = aimDic[key] + "\n" + value
        return aimDic

    def startAnalysis(self):
        xmlList, resultDicList = self.convertEvtxToXmlList(self.filePath)
        self.writeDicListToFile(resultDicList)

    def writeDicListToFile(self, resultDicList):
        print("开始导出文件")
        wb = oxl.Workbook()
        ws = wb.active
        ws.title = f"{self.fileName} 解析结果"
        self.writeExcellHead(ws, ["序号"] + self.requireTagList)

        for index, nowResultDic in enumerate(resultDicList):
            print(f"\r正在写入{index + 1}/{len(resultDicList)}行", end="")
            self.writeExcellCell(ws, index + 2, 1, str(index + 1), 0, True)
            for colIndex, nowKey in enumerate(self.requireTagList):
                self.writeExcellCell(
                    ws,
                    index + 2,
                    colIndex + 2,
                    (
                        " "
                        if str(nowResultDic[nowKey]) == ""
                        else str(nowResultDic[nowKey])
                    ),
                    0,
                    True,
                )
            self.writeExcellSpaceCell(ws, index + 2, len(self.requireTagList) + 2)

        colWidthArr = [10] + [20] * len(self.requireTagList)
        self.setExcellColWidth(ws, colWidthArr)

        fileName = f"{self.fileName}"
        filePath = os.path.join(self.outputDir, f"{fileName}.xlsx")
        self.saveExcell(wb, filePath)
        print(f"<--成功导出文件：{filePath}")

    # Excel utility methods
    def getExcellStyleDic(self):
        styleDic = {}
        styleDic["thin"] = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        styleDic["align"] = Alignment(horizontal="center", vertical="center")
        styleDic["left"] = Alignment(horizontal="left", vertical="center")
        styleDic["right"] = Alignment(horizontal="right", vertical="center")
        styleDic["bold"] = Font(bold=True)
        styleDic["link"] = Font(color="0000FF")
        styleDic["underLine"] = Font(underline="single")
        return styleDic

    def writeExcellHead(self, ws, headArr):
        styleDic = self.getExcellStyleDic()
        for index, head in enumerate(headArr):
            ws.cell(row=1, column=index + 1).value = head
            ws.cell(row=1, column=index + 1).border = styleDic["thin"]
            ws.cell(row=1, column=index + 1).alignment = styleDic["align"]
            ws.cell(row=1, column=index + 1).font = styleDic["bold"]
        return ws

    def writeExcellCell(
        self,
        ws,
        row,
        column,
        value,
        borderNum,
        ifAlign,
        hyperLink=None,
        fgColor="FFFFFF",
        otherAlign=None,
    ):
        styleDic = self.getExcellStyleDic()
        aimCell = ws.cell(row=row, column=column)
        aimCell.value = value

        if borderNum == 0:
            aimCell.border = styleDic["thin"]

        if ifAlign:
            aimCell.alignment = styleDic["align"]
        elif otherAlign is not None:
            aimCell.alignment = (
                styleDic["left"] if otherAlign == 0 else styleDic["right"]
            )

        if hyperLink:
            aimCell.hyperlink = hyperLink
            aimCell.font = styleDic["link"]

        fill = PatternFill("solid", fgColor=fgColor)
        aimCell.fill = fill
        return ws

    def writeExcellSpaceCell(self, ws, row, column):
        ws.cell(row=row, column=column).value = " "
        return ws

    def setExcellColWidth(self, ws, colWidthArr):
        for colWidindex, width in enumerate(colWidthArr):
            ws.column_dimensions[chr(ord("A") + colWidindex)].width = width
        return ws

    def saveExcell(self, wb, savePath):
        if os.path.exists(savePath):
            os.remove(savePath)
        wb.save(savePath)
        return True


def process_folder(folder_path):
    output_dir = os.path.join(folder_path, "output")
    os.makedirs(output_dir, exist_ok=True)

    evtx_files = glob.glob(os.path.join(folder_path, "*.evtx"))

    for evtx_file in evtx_files:
        print(f"Processing file: {evtx_file}")
        try:
            AnalysisEvtx(evtx_file, output_dir)
        except Exception as ex:
            print(f"处理文件 {evtx_file} 时发生异常：{str(ex)}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python main.py <folder_path>")
        sys.exit(1)

    folder_path = sys.argv[1]
    if not os.path.isdir(folder_path):
        print(f"Error: {folder_path} is not a valid directory")
        sys.exit(1)

    process_folder(folder_path)
